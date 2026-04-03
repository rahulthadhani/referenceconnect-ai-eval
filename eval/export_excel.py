import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

Path("results").mkdir(exist_ok=True)

# Load data
v1 = pd.read_csv("results/eval_v1.csv")
v2 = pd.read_csv("results/eval_v2.csv")

# Build summary by category
cat_v1 = v1.groupby("category").agg(
    relevance_v1=("relevance_score", "mean"),
    accuracy_v1=("keyword_accuracy", "mean"),
    latency_v1=("latency_ms", "mean"),
    word_count_v1=("word_count", "mean")
).round(2)

cat_v2 = v2.groupby("category").agg(
    relevance_v2=("relevance_score", "mean"),
    accuracy_v2=("keyword_accuracy", "mean"),
    latency_v2=("latency_ms", "mean"),
    word_count_v2=("word_count", "mean")
).round(2)

summary = cat_v1.join(cat_v2)
summary["relevance_delta"] = (summary["relevance_v2"] - summary["relevance_v1"]).round(2)
summary["accuracy_delta"]  = (summary["accuracy_v2"]  - summary["accuracy_v1"]).round(2)
summary = summary.reset_index()

# Overall summary row
overall = pd.DataFrame([{
    "category": "OVERALL",
    "relevance_v1": v1["relevance_score"].mean().round(2),
    "accuracy_v1":  v1["keyword_accuracy"].mean().round(2),
    "latency_v1":   v1["latency_ms"].mean().round(2),
    "word_count_v1": v1["word_count"].mean().round(2),
    "relevance_v2": v2["relevance_score"].mean().round(2),
    "accuracy_v2":  v2["keyword_accuracy"].mean().round(2),
    "latency_v2":   v2["latency_ms"].mean().round(2),
    "word_count_v2": v2["word_count"].mean().round(2),
    "relevance_delta": (v2["relevance_score"].mean() - v1["relevance_score"].mean()).round(2),
    "accuracy_delta":  (v2["keyword_accuracy"].mean() - v1["keyword_accuracy"].mean()).round(2),
}])
summary = pd.concat([summary, overall], ignore_index=True)

# Merge all results
all_results = pd.concat([v1, v2], ignore_index=True).sort_values(["query_id","model_version"])

# Regression table
merged = v1[["query_id","query","category","relevance_score"]].merge(
    v2[["query_id","relevance_score"]], on="query_id", suffixes=("_v1","_v2")
)
merged["delta"] = merged["relevance_score_v2"] - merged["relevance_score_v1"]
regressed = merged[merged["delta"] < 0].reset_index(drop=True)

# Write to Excel
output_path = "results/eval_report.xlsx"
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    summary.to_excel(writer,     sheet_name="Summary by Category", index=False)
    all_results.to_excel(writer, sheet_name="All Results",          index=False)
    v1.to_excel(writer,          sheet_name="V1 Baseline",          index=False)
    v2.to_excel(writer,          sheet_name="V2 Improved",          index=False)
    if len(regressed) > 0:
        regressed.to_excel(writer, sheet_name="Regressions",        index=False)

# Style the Summary sheet
wb = load_workbook(output_path)
ws = wb["Summary by Category"]

# Colors
BLUE_DARK  = "1d4ed8"
BLUE_LIGHT = "dbeafe"
GREEN      = "dcfce7"
RED        = "fee2e2"
GRAY       = "f1f5f9"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor=BLUE_DARK)
center      = Alignment(horizontal="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

# Style headers
for cell in ws[1]:
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center
    cell.border    = thin_border

# Style data rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border    = thin_border
        cell.alignment = center
        # Highlight OVERALL row
        if cell.row == ws.max_row:
            cell.fill = PatternFill("solid", fgColor=BLUE_LIGHT)
            cell.font = Font(bold=True, size=11)

# Color delta columns green/red
delta_cols = [
    ws.max_column - 1,  # relevance_delta
    ws.max_column       # accuracy_delta
]
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for col_idx in delta_cols:
        cell = row[col_idx - 1]
        if isinstance(cell.value, (int, float)):
            if cell.value > 0:
                cell.fill = PatternFill("solid", fgColor=GREEN)
            elif cell.value < 0:
                cell.fill = PatternFill("solid", fgColor=RED)

# Auto-width columns
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

wb.save(output_path)
print(f"[OK] Saved {output_path}")